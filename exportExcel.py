import os
from DBConnect import mysql_connection
from openpyxl import Workbook
from dotenv import load_dotenv


class exportExcel:
    """Add doc string"""
    def __init__(self) -> None:
        self._load_env_var = load_dotenv()
        self.config = {
        'user': os.getenv('USERNAME'),
        'password': os.getenv('PASSWORD'),
        'host': os.getenv('HOST'),
        'database': os.getenv('DATABASE')
        }
        self._check_config()
        self._check_connection()

    def _check_config(self)-> None:
        """ function checks to see if credentials are missing 
            or not present in env variables"""
        for key in self.config.keys():
            if self.config[key] == None:
                self.config[key] = input(f"Enter Database {key}: ")
    
    def _check_connection(self) -> None:
        """ function checks to see if connection is viable with provided credintials"""
        pass

    def exportMysqlToExcelWorkbook(self, filepath = ".") -> None:
        """Add Doc String"""
            
        with mysql_connection(**self.config) as conn:
            with conn.cursor() as cursor:
                cursor.execute("SHOW TABLES;")
                tables = {i[0]:[] for i in cursor }

                for table in tables:
                    cursor.execute(f"SHOW columns FROM {table}")
                    tables[table] = [column[0] for column in cursor.fetchall()]

        # create excel workbook
        wb = Workbook()

        for table in tables.keys():
            _sheet = wb.create_sheet(title=table)
            _sheet.append(tables[table])

        del wb['Sheet']
        file = f"{filepath}/{self.config['database']}.xlsx"
        wb.save(filename=file)
        print(f"Database has been exported to {file}")

    def exportSqlservertoExcelWorkbook(self) -> None:
        "Add Docstring"
        pass

    def exportSqliteToExcelWorkbook(self) -> None:
        "Add Docstring"
        pass


if __name__ == "__main__":
    db = exportExcel()
    db.exportMysqlToExcelWorkbook()
